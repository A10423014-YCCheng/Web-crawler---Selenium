from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
import requests
from openpyxl.drawing.image import Image
from openpyxl.workbook import Workbook

s = Service(r"C:\chromedriver.exe")
driver = webdriver.Chrome(service=s)

###進入天瓏並搜尋selenium###
url = 'https://www.tenlong.com.tw/'
driver.get(url)

xpathstr = '/html/body/div[3]/nav[1]/nav/div/form/input[2]'  # Copy full Xpath取得
search_field = driver.find_element(By.XPATH, xpathstr)  # 使用By.XPATH方式
search_field.send_keys('selenium')  # 輸入selenium到對應位置
search_field.submit()  # 提交

# 搜尋selenium完的畫面#
###取得書名###
book_name = driver.find_elements(By.CLASS_NAME, 'book-data')

book_l = []
for item in book_name:
    book_l.append(item.text)

title_l = []
for item in book_l:
    title_l.append(item.split('\n')[0])

###取得每本書的連結###
link_l = []
for title in title_l:
    book_link = driver.find_elements(By.LINK_TEXT, title)
    for item in book_link:
        link_l.append(item.get_attribute('href'))

###進去每本書的網頁，抓圖片url&書的完整名稱###
complete_title = []
img_url = []
for link in link_l:
    driver.get(link)
    url = driver.find_element(By.XPATH, '/html/body/div[3]/div[3]/div/div[1]/article/section[1]/div[1]/div[2]/div/a[1]')
    img_url.append(url.get_attribute('data-featherlight'))
    name = driver.find_element(By.CLASS_NAME, 'item-title')
    complete_title.append(name.text)
    time.sleep(3)

###抓圖片&存檔###
# 存檔路徑
path = r'D:\Python\Program\DB and Web crawler\20221116\Image\\'


# 去掉書名中不能存成檔名的符號
def filename_replace(str1):
    str1 = str1.replace('/', '').replace('\\', '').replace(':', '').replace('?', '').replace('*', '').replace('"','').replace('>', '').replace('<', '').replace('|', '')
    return str1


for i, item in enumerate(img_url):
    r = requests.get(item)  # 到圖片網頁去
    with open(path + filename_replace(complete_title[i]) + '.jpg', 'wb') as f:
        f.write(r.content)  # 存檔
driver.quit()

###存到Excel###
wb = Workbook()
ws1 = wb.create_sheet("Selenium")

ws1.column_dimensions['A'].width = 15
wb = Workbook()
ws1 = wb.create_sheet("Selenium")
ws1.column_dimensions['A'].width = 15
for i,item in enumerate(complete_title):
    ws1.cell(row=i + 1, column=2, value=item)
    x = i+1
    ws1.row_dimensions[x].height = 105
    img = Image(path + filename_replace(item) + '.jpg')
    imgsize = (1997/20,2572/20)
    img.width, img.height = imgsize
    position = 'A' + str(x)
    ws1.add_image(img,position)
wb.save(r"D:\Python\Program\DB and Web crawler\20221116\Selenium.xlsx")